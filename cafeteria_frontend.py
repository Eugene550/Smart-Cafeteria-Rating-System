"""
Smart Cafeteria Rating System — Stage 1 (front end), CSV-ready version
Google Form CSV  ->  per-review aspect scores in [0, 1]

Division of labour:
  - Lexicon (Variable_and_keywords_refined.xlsx): WHICH aspect a word is about + DIRECTION (Level).
  - VADER: the STRENGTH (number), with negation/intensifier handling.
  - DOMAIN_EXTRA below: a few cafeteria-specific words VADER/the sheet miss
    (flies, stomach ache, packed...). To be confirmed/expanded with the expert.
"""

import re
import pandas as pd
from openpyxl import load_workbook
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

ASPECTS = ["Food Quality", "Cleanliness", "Menu Variety", "Service Speed"]
LEVEL_VALENCE = {"High": 2.5, "Medium": 0.3, "Low": -2.5}

GENERIC = {"okay", "ok", "fine", "average", "moderate", "normal",
           "acceptable", "good", "nice", "bad", "not bad"}

ASPECT_TRIGGERS = {
    "Food Quality":  ["food", "meal", "dish", "taste", "flavour", "flavor", "portion", "cuisine"],
    "Cleanliness":   ["clean", "cleanliness", "hygiene", "toilet", "floor", "table"],
    "Menu Variety":  ["menu", "choice", "selection", "option", "variety"],
    "Service Speed": ["service", "queue", "wait", "line", "staff", "counter", "speed"],
}

# Extra domain words VADER doesn't know (seen in the real reviews). word -> level.
# These also act as aspect triggers, so "flies" now flags Cleanliness.
DOMAIN_EXTRA = {
    "Cleanliness":   [("flies", "Low"), ("fly", "Low"), ("cockroach", "Low"),
                      ("roach", "Low"), ("pest", "Low"), ("insect", "Low"),
                      ("smelly", "Low"), ("stinky", "Low")],
    "Food Quality":  [("expired", "Low"), ("undercooked", "Low"), ("raw", "Low"),
                      ("oily", "Low"), ("greasy", "Low"),
                      ("stomach ache", "Low"), ("stomachache", "Low")],
    "Service Speed": [("packed", "Low"), ("jam", "Low")],
    "Menu Variety":  [],
}


def load_lexicon(path):
    wb = load_workbook(path, read_only=True)
    ws = wb["Lexicon"]
    matching = {a: set(ASPECT_TRIGGERS[a]) for a in ASPECTS}
    vader_updates = {}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # add the expert lexicon rows
    for row in rows:
        var, kw, lvl = row[0], row[1], row[2]
        if var is None:
            continue
        kw = str(kw).strip().lower()
        vader_updates[kw.replace(" ", "_")] = LEVEL_VALENCE.get(lvl, 0.0)
        if kw not in GENERIC:
            matching[var].add(kw)
    # add the extra domain words
    for aspect, extras in DOMAIN_EXTRA.items():
        for kw, lvl in extras:
            kw = kw.lower()
            vader_updates[kw.replace(" ", "_")] = LEVEL_VALENCE.get(lvl, 0.0)
            matching[aspect].add(kw)
    matching = {a: sorted(w, key=len, reverse=True) for a, w in matching.items()}
    return matching, vader_updates


def build_analyzer(vader_updates):
    sia = SentimentIntensityAnalyzer()
    sia.lexicon.update(vader_updates)
    return sia


def split_clauses(text):
    # normalise unicode punctuation and newlines, then split into clauses
    text = (str(text).lower()
            .replace("，", ",").replace("。", ".").replace("、", ",")
            .replace("\n", " . "))
    parts = re.split(r"[.;,!?/]| but | however | though | and | while ", text)
    return [p.strip() for p in parts if p.strip()]


def detect_aspects(clause, matching):
    return {a for a, words in matching.items() if any(w in clause for w in words)}


def _collapse_multiword(clause, vu):
    for token in vu:
        if "_" in token:
            clause = clause.replace(token.replace("_", " "), token)
    return clause


def score_review(review, sia, matching, vu):
    per = {a: [] for a in ASPECTS}
    for clause in split_clauses(review):
        asp = detect_aspects(clause, matching)
        if not asp:
            continue
        c = _collapse_multiword(clause, vu)
        s = (sia.polarity_scores(c)["compound"] + 1) / 2
        for a in asp:
            per[a].append(s)
    return {a: round(sum(v) / len(v), 3) for a, v in per.items() if v}


def load_reviews_csv(path):
    raw = pd.read_csv(path)
    df = raw.iloc[:, :4].copy()
    df.columns = ["timestamp", "occupation", "cafeteria", "review"]
    if "review_clean" in raw.columns:           # carry cleaned text if the file has it
        df["review_clean"] = raw["review_clean"]
    return df.dropna(subset=["review"]).reset_index(drop=True)


def score_dataframe(df, sia, matching, vu, text_col=None):
    """Add one score column per aspect (NaN where the review didn't mention it).
    Uses the 'review_clean' column automatically if present, else 'review'."""
    if text_col is None:
        text_col = "review_clean" if "review_clean" in df.columns else "review"
    out = df.copy()
    scores = df[text_col].apply(lambda r: score_review(r, sia, matching, vu))
    for a in ASPECTS:
        out[a] = scores.apply(lambda d: d.get(a))
    return out


if __name__ == "__main__":
    LEX = "Variable_and_keywords_refined.xlsx"
    CSV = "reviews_clean.csv"

    matching, vu = load_lexicon(LEX)
    sia = build_analyzer(vu)
    df = load_reviews_csv(CSV)
    scored = score_dataframe(df, sia, matching, vu)

    print(f"Loaded {len(df)} reviews across {df['cafeteria'].nunique()} cafeterias\n")
    print("Reviews per cafeteria:")
    print(df["cafeteria"].value_counts().to_string(), "\n")

    print("Coverage (non-missing scores per aspect):")
    for a in ASPECTS:
        print(f"  {a}: {scored[a].notna().sum()}/{len(scored)}")

    print("\nSample scored reviews:")
    for i in range(6):
        row = scored.iloc[i]
        vals = {a: row[a] for a in ASPECTS if pd.notna(row[a])}
        print(f"  [{i}] {row['cafeteria'][:18]:18}  {vals}")
